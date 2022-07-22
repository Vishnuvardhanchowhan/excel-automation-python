import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
for row in range(1,sheet.max_row+1):
    cell1=sheet.cell(row,3)
    correct_price = cell1.value * 3
    corrected_price_cell=sheet.cell(row,4)
    corrected_price_cell.value=correct_price
values = Reference(sheet,min_row=2,max_row=sheet.max_row
                   ,min_col=4,max_col=4)
chart =BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save("transactions2.xlsx")
